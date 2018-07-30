import json
from openpyxl import load_workbook
from openpyxl.styles import Font
from project_specifications.models import Specification
from warehouse.models import GoodKind, Unit, Manufacturer
from synergycrm.exceptions import SngyException
from os.path import join
import datetime
import re


def xstr(s):
	if s is None:
		return ''
	if isinstance(s, float):
		return ('%.2f' % s).replace('.', ',')
	return str(s)


class MyFileInput:
	def __init__(self, all_data, dir_input, output_dir):
		all_data = all_data
		self.book = load_workbook(join(dir_input, 'input.xlsx'), data_only=True)
		self.sheet = self.book[self.book.sheetnames[1]]
		if len(all_data) < 1:
			raise SngyException('Нельзя экспортировать пустую спецификацию')
		specification = list(Specification.objects.filter(id=all_data[0]['specification_id']).values())[0]
		MySpecificationLegend(specification, self.sheet)

		self.read_input(all_data)
		now = datetime.datetime.now()
		filename = '%s_от_%s.xlsx' % (specification['pressmark'].replace('/', '_'), now.strftime("%d-%m-%y_%H-%M"))
		self.filename = filename
		self.book.save(join(output_dir, filename))

	# Метод, читающий входной файл
	def read_input(self, all_data):
		# Функция генерации названия позиций и сохранения сгруппированных строк
		def generate_pos(saved_row, all_unit_id, designation_data, start_counter, start_text, counter):
			designations = [p for p in str(saved_row['positional_designation']).split(',') if p.strip()]
			saved_row['positional_designation'] = ','.join(designations)
			if len(designations) > 1:
				for index, designation in enumerate(designations):
					# Строка positional_designation разбивается на 2 части: text и num
					num = r.findall(designation)[-1]
					num_len = len(num)
					text = designation[:len(designation)-num_len]
					if not counter:
						counter = int(num)
						start_counter = int(num)
						start_text = text
					else:
						# Последовательность: Если текущий элемент имеет то же имя и заканчивается на единицу больше
						if int(num) == counter+1 and text == start_text:
							counter += 1
						# Окончание последовательности
						if not int(num) == counter or not text == start_text or index+1 == len(designations):
							if start_counter+1 < counter:
								designation_data += start_text + str(start_counter) + '...' + str(counter)
							else:
								designation_data += start_text + str(start_counter)
								for item in range(start_counter+1, counter+1):
									designation_data += ', ' + str(item)

							# Проверка Если ПОСЛЕДНИЙ элемент списка текущего GoodKind'а выбивается из последовательности
							if index+1 == len(designations):
								if not int(num) == counter or not text == start_text:
									designation_data += ', ' + text + str(num)

							counter = int(num)
							start_counter = int(num)
							start_text = text
							if not index+1 == len(designations):
								designation_data += ', '
				row = {
					'id': saved_row['id'],
					'good_kind_id': saved_row['good_kind_id'],
					'description_info': saved_row['description_info'],
					'unit_id': saved_row['unit_id'],
					'specification_id': saved_row['specification_id'],
					'positional_designation': designation_data,
					'count': saved_row['count'],
					'position_in_table': saved_row['position_in_table'],
					'grouping_name': saved_row['grouping_name'],
					'note': saved_row['note']
				}
			else:
				row = {
					'id': saved_row['id'],
					'good_kind_id': saved_row['good_kind_id'],
					'description_info': saved_row['description_info'],
					'unit_id': saved_row['unit_id'],
					'specification_id': saved_row['specification_id'],
					'positional_designation': saved_row['positional_designation'],
					'count': saved_row['count'],
					'position_in_table': saved_row['position_in_table'],
					'grouping_name': saved_row['grouping_name'],
					'note': saved_row['note']
				}
			my_new_rows.append(row)
			return False

		ws = self.book[self.book.sheetnames[0]]
		main_font = Font(name='Calibri', size=11, underline='single')
		# all_data - это SpecificationsPositions
		# Отсортировано по полю position_in_table
		spec_rows = all_data
		my_new_rows = []
		saved_row = None
		# Счётчик. изменяется от начального числа пока не изменится имя
		counter = None
		# Финальная готовая строка
		designation_data = ''
		# РегВ возвращает числа в конце строки
		r = re.compile('.*?(\d+)')

		# Объединяем строки
		for spec_row in spec_rows:
			# saved_row - строка с прошлой итерации
			# Пропускает первую итерацию
			if saved_row is not None:
				if spec_row['good_kind_id'] == saved_row['good_kind_id'] and spec_row['good_kind_id'] is not None:
					saved_row['count'] += spec_row['count']
					saved_row['positional_designation'] += ',' + spec_row['positional_designation']
					# Записываем первый попавшийся комментарий
					if not saved_row['note']:
						saved_row['note'] = spec_row['note']
					# all_unit_id получаем у первого элемента каждого GoodKind'а
					# Единицы должны быть одинаковыми иначе ошибка
					if not all_unit_id == spec_row['unit_id']:
						raise SngyException('Ошибка: В одном виде товаров присутствуют разные Единицы измерения!')
					continue
				# saved_row
				# if saved_row['']
				generate_pos(saved_row, all_unit_id, designation_data, start_counter, start_text, counter)
			# Доходим до сюда если у строки Сменился GoodKind
			saved_row = spec_row
			all_unit_id = spec_row['unit_id']
			designation_data = ''
			counter = 0
			start_counter = 0
			start_text = ''

		# Заполняем в конце последние данные
		if saved_row:
			generate_pos(saved_row, all_unit_id, designation_data, start_counter, start_text, counter)

		for index, spec_row in enumerate(my_new_rows):
			if spec_row['grouping_name']:
				ws['I' + str(5 + index)].font = main_font
				ws['I' + str(5 + index)].value = spec_row['grouping_name']
				# Добавить подчёркивание
			else:
				good_kind = list(GoodKind.objects.filter(id=spec_row['good_kind_id']).values())[0]
				unit = list(Unit.objects.filter(id=spec_row['unit_id']).values())[0]
				manufacturer = list(Manufacturer.objects.filter(id=good_kind['manufacturer_id']).values())[0]
				if manufacturer['id'] == 52:
					manufacturer['name'] = ''
				if good_kind['code'] is None:
					good_kind['code'] = ''
				ws['H' + str(5 + index)].value = spec_row['positional_designation']
				ws['I' + str(5 + index)].value = good_kind['name']
				ws['J' + str(5 + index)].value = spec_row['description_info']
				ws['K' + str(5 + index)].value = good_kind['code']
				ws['L' + str(5 + index)].value = manufacturer['name']
				ws['M' + str(5 + index)].value = unit['short_name']
				ws['N' + str(5 + index)].value = spec_row['count']
				ws['O' + str(5 + index)].value = good_kind['mass']
				ws['P' + str(5 + index)].value = spec_row['note']


class MySpecificationLegend:
	def __init__(self, specification, ws):
		ws['H2'].value = specification['pressmark']
		ws['H4'].value = specification['object_name']
		ws['H7'].value = specification['section_name']
		ws['H10'].value = specification['document_name']
		ws['O8'].value = specification['state']
		ws['P8'].value = '1'
		ws['O10'].value = specification['organization']
		posts = ['Разработал', 'Проверил', 'Н. контр.', 'ГИП']
		names = specification['workers_data']
		names = json.loads(names)
		new_names = []
		for q in names:
			if q['name']:
				new_names.append(q['name'].split(' ')[0])
			else:
				new_names.append('')
		names = new_names
		i = 0
		for row in range(7, 9):
			if i >= len(names):
				break
			ws['B' + str(row)].value = posts[i]
			ws['D' + str(row)].value = names[i]
			if specification['dates'][i] is not None:
				date = list(specification['dates'][i])
				if len(date) == 4:
					date_string = date[0] + date[1] + '.' + date[2] + date[3]
					ws['G' + str(row)].value = str(date_string)
				elif len(date) == 0:
					ws['G' + str(row)].value = ''
				else:
					ws['G' + str(row)].value = specification['dates'][i]
			else:
				ws['G' + str(row)].value = specification['dates'][i]
			i += 1

		for row in range(11, 13):
			if i >= len(names):
				break
			ws['B' + str(row)].value = posts[i]
			ws['D' + str(row)].value = names[i]
			if specification['dates'][i] is not None:
				date = list(specification['dates'][i])
				if len(date) == 4:
					date_string = date[0] + date[1] + '.' + date[2] + date[3]
					ws['G' + str(row)].value = str(date_string)
				elif len(date) == 0:
					ws['G' + str(row)].value = ''
				else:
					ws['G' + str(row)].value = specification['dates'][i]
			else:
				ws['G' + str(row)].value = specification['dates'][i]
			i += 1
